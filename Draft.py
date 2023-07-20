import random, os
import PySimpleGUI as sg
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys


# create the brand class
class Brand:
    # constructor
    def __init__(self, name, men, women, title_names):
        self.name = name
        self.men = men
        self.women = women
        self.roster = men + women
        self.title_names = title_names

    # method to create tag teams
    def createTeams(self, team_number):
        # create variables
        men = self.men
        self.teams = []

        random.shuffle(men)

        # assign teams
        for i in range(0, team_number * 2, 2):
            self.teams.append((men[i], men[i + 1]))
        return self.teams

    def createWomenTeams(self, womens_team_number):
        women = self.women
        self.women_teams = []
        if womens_team_number == 0: return self.women_teams

        random.shuffle(women)

        # assign teams
        for i in range(0, womens_team_number * 2, 2):
            self.women_teams.append((women[i], women[i + 1]))
        return self.women_teams

    # method to assign champions
    def assignChampions(self, teams, second_midcard, womens_midcard):
        male_roster = self.men
        female_roster = self.women

        # randomize order of sections
        random.shuffle(male_roster)
        random.shuffle(female_roster)
        random.shuffle(teams)

        # create variables
        self.champions = []
        second_midcard_champion = None
        womens_midcard_champion = None

        singles_division = Brand.removeTeams(male_roster, teams)

        # assign champions
        world_champion = singles_division[0]
        midcard_champion = singles_division[1]
        womens_champion = female_roster[0]
        tag_champions = teams[0]
        if second_midcard: second_midcard_champion = singles_division[2]
        if womens_midcard: womens_midcard_champion = female_roster[1]

        # put champions in list
        self.champions = [world_champion, midcard_champion, womens_champion, tag_champions,
                          second_midcard_champion, womens_midcard_champion]
        return self.champions

    # method to assign divisions
    def assignDivisions(self, tag_teams, champions, second_midcard, womens_midcard):
        # create the variables
        men = self.men
        women = self.women
        self.divisions = []
        world_division, midcard_division, womens_division, tag_division, second_midcard_division, womens_midcard_division = [], [], [], [], [], []
        male_belt_order = 1
        female_belt_order = 1

        # prepare lists for divisions
        Brand.removeTeams(men, tag_teams)

        # shuffle the lists
        random.shuffle(men)
        random.shuffle(women)
        random.shuffle(tag_teams)

        # Create the men's divisions
        for m in men:

            # World Title
            if male_belt_order == 1:
                world_division.append(m)
                male_belt_order = 2

            # Midcard Title
            elif male_belt_order == 2:
                midcard_division.append(m)
                if second_midcard:
                    male_belt_order = 3
                else:
                    male_belt_order = 1

            # 2nd Midcard Title
            elif male_belt_order == 3:
                second_midcard_division.append(m)
                male_belt_order = 1

        # Women's Division(s)
        for w in women:
            if female_belt_order == 1:
                womens_division.append(w)
                if womens_midcard: female_belt_order = 2

            elif female_belt_order == 2:
                womens_midcard_division.append(w)
                female_belt_order = 1
        tag_division = tag_teams

        self.divisions = [world_division, midcard_division, womens_division, tag_division,
                          second_midcard_division, womens_midcard_division]
        return self.divisions

    def assignFeuds(self, divisions):
        # create the variables
        self.feuds = []
        world_div, mid_div, women_div, tag_div = divisions[0], divisions[1], divisions[2], divisions[3]
        random.shuffle(world_div);
        random.shuffle(mid_div);
        random.shuffle(tag_div);
        random.shuffle(women_div)
        self.feuds.append(f"{world_div[0]} Vs. {world_div[1]}")
        self.feuds.append(f"{mid_div[0]} Vs. {mid_div[1]}")
        self.feuds.append(f"{women_div[0]} Vs. {women_div[1]}")
        self.feuds.append(f"{tag_div[0][0]} & {tag_div[0][1]} Vs. {tag_div[1][0]} & {tag_div[1][1]}")
        return self.feuds

    @staticmethod
    def removeTeams(roster, teams):
        singles_roster = roster
        for r in singles_roster:
            for t in teams:
                if r == t[0] or r == t[1]:
                    singles_roster.remove(r)
        return singles_roster

# additional functions

# create the rosters
def getRoster(roster_file):
    with open(roster_file) as f:
        roster = f.readlines()
    return roster


# draft the brands
def draftRoster(roster, brands):
    raw_roster, smackdown_roster, nxt_roster, aew_roster, nxt_uk_roster, roh_roster = [], [], [], [], [], []  # Roster lists
    b = 1  # Brand Counter
    random.shuffle(roster)  # Shuffle the roster order

    # Draft the rosters
    for r in roster:

        # Raw Pick
        if b == 1:
            raw_roster.append(r[:-1])
            if brands > 1: b = 2
            continue

        # Smackdown Pick
        if b == 2:
            smackdown_roster.append(r[:-1])
            if brands == 2:
                b = 1
                continue
            else:
                b = 3
                continue

        # NXT Pick
        if b == 3:
            nxt_roster.append(r[:-1])
            if brands == 3:
                b = 1
                continue
            else:
                b = 4
                continue

        # AEW Pick
        if b == 4:
            aew_roster.append(r[:-1])
            if brands == 4:
                b = 1
                continue
            else:
                b = 5
                continue

        # NXT UK Pick
        if b == 5:
            nxt_uk_roster.append(r[:-1])
            if brands == 5:
                b = 1
                continue
            else:
                b = 6
                continue

        # ROH Pick
        if b == 6:
            roh_roster.append(r[:-1])
            b = 1

            continue

    brandRosters = [raw_roster, smackdown_roster, nxt_roster, aew_roster, nxt_uk_roster, roh_roster]
    return brandRosters


def addTeams(roster, teams):
    for t in teams:
        roster.append(t[0])
        roster.append(t[1])
    return roster


def makeTabLayout(brand_name, brand_key, title_names, key_names):
    brand_layout = [[sg.Text("Brand Name: "), sg.Input(key=brand_key, default_text=brand_name, enable_events=True)],
                    [sg.Text("World Title Name: "), sg.Input(key=key_names[0], default_text=title_names[0])],
                    [sg.Text("Midcard Title Name: "), sg.Input(key=key_names[1], default_text=title_names[1])],
                    [sg.Text("Women's Title Name: "), sg.Input(key=key_names[2], default_text=title_names[2])],
                    [sg.Text("Tag Team Title Name: "), sg.Input(key=key_names[3], default_text=title_names[3])],
                    [sg.Text("Second Midcard Title Name: "), sg.Input(key=key_names[4], default_text=title_names[4])],
                    [sg.Text("Women's Midcard Title Name: "), sg.Input(key=key_names[5], default_text=title_names[5])]
                    ]
    return brand_layout


# add the info of a brand to the text file
def makeTxtFile(f, brand, men, women, teams, champions, divisions, second_midcard_division, womens_midcard_division,
                feuds, title_names):
    # Header
    f.write(brand + "\n")
    f.write("=" * 30 + "\n\n")

    full_mens_roster = addTeams(men, teams)
    full_mens_roster.sort()
    women.sort()

    # Men's Roster
    f.write("Men's Roster\n")
    f.write("=" * 30 + "\n")
    for m in full_mens_roster:
        f.write(m + "\n")
    f.write("\n")

    # Women's Roster
    f.write("Women's Roster\n")
    f.write("=" * 30 + "\n")
    for w in women:
        f.write(w + "\n")
    f.write("\n")

    # Tag Teams
    f.write("" + "Tag Teams\n")
    f.write("=" * 30 + "\n")
    for t in teams:
        f.write(t[0] + " & " + t[1] + "\n")
    f.write("\n")

    # Champions
    f.write("Champions\n")
    f.write("=" * 30 + "\n")
    f.write(f"{title_names[0]} Champion: {champions[0]}\n") # World Title
    f.write(f"{title_names[1]} Champion: {champions[1]}\n") # Midcard Title
    if second_midcard_division: f.write(f"{title_names[4]} Champion: {champions[4]}\n")
    f.write(f"{title_names[3]} Champions: {champions[3][0]} & {champions[3][1]}\n") # Tag Team Title
    f.write(f"{title_names[2]} Champion: {champions[2]}\n")
    if womens_midcard_division: f.write(f"{title_names[5]} Champion: {champions[5]}")
    f.write("\n\n")

    # Divisions
    f.write("Divisions\n")
    f.write("=" * 30 + "\n\n")

    # World Title
    f.write(f"{title_names[0]} Title\n")
    f.write("=" * 30 + "\n")
    for d in divisions[0]:
        f.write(d + "\n")
    f.write("\n\n")

    # Midcard Title
    f.write(f"{title_names[1]} Title\n")
    f.write("=" * 30 + "\n")
    for d in divisions[1]:
        f.write(d + "\n")
    f.write("\n\n")

    # 2nd Midcard Title
    if second_midcard_division:
        f.write("2nd Midcard Title\n")
        f.write("=" * 30 + "\n")
        for d in divisions[4]:
            f.write(f"{d}\n")
        f.write("\n\n")

    # Tag Title
    f.write(f"{title_names[3]} Title\n")
    f.write("=" * 30 + "\n")
    for d in divisions[3]:
        f.write(d[0] + " & " + d[1] + "\n")
    f.write("\n\n")

    # Women's Title
    f.write(f"{title_names[2]} Title\n")
    f.write("=" * 30 + "\n")
    for d in divisions[2]:
        f.write(d + "\n")
    f.write("\n\n")

    # Women's Midcard Title
    if womens_midcard_division:
        f.write("Women's Midcard Title\n")
        f.write("=" * 30 + "\n")
        for d in divisions[5]:
            f.write(f"{d}\n")
    f.write("\n\n")

    # Feuds
    f.write("Rivalries\n")
    f.write("=" * 30 + "\n")
    for r in feuds:
        f.write(f"{r}\n")
    f.write("\n\n")


# add the women's tag division to the text file
def writeWomenTag(f, womens_teams, womens_tag_champions):
    f.write("\n\n")
    f.write("Women's Tag Division\n")
    f.write("=" * 30)
    f.write("\n\n")
    f.write("Women\'s Tag Team Champions: {} & {}".format(womens_tag_champions[0], womens_tag_champions[1]))
    f.write("\n\n")
    for w in womens_teams[1:]:
        f.write(w[0] + " & " + w[1] + "\n")


# graphical interface of the program
def gui():
    current_directory = os.getcwd()
    sg.theme("Reddit")

    # creates list of default championship names
    title_names = [
        ["WWE", "United States", "Raw Tag Team", "Raw Women's", "24/7", "Women's United States"], # Raw
        ["Universal", "Intercontinental", "Smackdown Women's", "Smackdown Tag Team", "European", "Women's Intercontinental"], # Smackdown
        ["NXT", "North American", "NXT Women's", "NXT Tag Team", "Cruiserweight", "Women's North American"], # NXT
        ["AEW World", "TNT", "AEW Women's", "AEW Tag Team", "All Atlantic", "TBS"], # AEW
        ["NXT UK", "Heritage Cup", "NXT UK Women's", "NXT UK Tag Team", "NXT UK Cruiserweight", "Women's Heritage Cup"], # NXT UK
        ["ROH World", "Television", "ROH Women's", "ROH Tag Team", "Pure", "Women's Television"] # ROH
    ]

    # creates list of key names
    key_names = [
        ["raw_world", "raw_mid", "raw_women", "raw_tag", "raw_mid2", "raw_womens_mid"], # Raw
        ["sd_world", "sd_mid", "sd_women", "sd_tag", "sd_mid2", "sd_womens_mid"], # Smackdown
        ["nxt_world", "nxt_mid", "nxt_women", "nxt_tag", "nxt_mid2", "nxt_womens_mid"], # NXT
        ["aew_world", "aew_mid", "aew_women", "aew_tag", "aew_mid2", "aew_womens_mid"], # AEW
        ["uk_world", "uk_mid", "uk_women", "uk_tag", "uk_mid2", "uk_womens_mid"], # NXT UK
        ["roh_world", "roh_mid", "roh_women", "roh_tag", "roh_mid2", "roh_womens_mid"] # ROH
    ]

    # handle case when excel file is already open
    try:
        draft_file = open("Draft.xlsx", 'w')
        draft_file.close()
        draft_file = open("Draft.xlsx", 'a')
    except IOError:
        sg.PopupError("The Draft.xlsx file is already open.\nPlease close the file before running this program")
        sys.exit()

    # GUI Variables
    men = sg.FileBrowse("Men's Roster", file_types=[("TXT Files", "*.txt")], initial_folder=current_directory)
    women = sg.FileBrowse("Women's Roster", file_types=[("TXT Files", "*.txt")], initial_folder=current_directory)

    brands = [i for i in range(1, 7)]
    teams = [i for i in range(1, 21)]
    women_teams = [i for i in range(0, 11)]

    # GUI Layout
    main_layout = [
        [sg.Text("Select Your Male and Female Roster Files")],  # Title
        [sg.InputText(key="-FILE_PATH-"), men],  # Men's Roster Input
        [sg.InputText(key="-FILE_PATH2-"), women],  # Women's Roster Input
        [sg.Text("Select the number of brands"),  # Brand Number
         sg.Spin(brands, initial_value=2, key='brands', size=4, enable_events=True)],
        [sg.Text("Select the number of tag teams per brand"),  # Team Number
         sg.Spin(teams, key='teams', initial_value=4, size=4)],
        [sg.Text("Select the number of Women's Tag Teams you want per brand"),  # Women's Team Number
         sg.Spin(women_teams, initial_value="0", key='wTag', size=4)],
        [sg.Checkbox("2nd Midcard Title", key='mid2')],
        [sg.Checkbox("Women's Midcard Title", key='wMid')]
    ]

    raw_layout = makeTabLayout("Raw", "raw_name", title_names[0], key_names[0])
    sd_layout = makeTabLayout("Smackdown", "sd_name", title_names[1], key_names[1])
    nxt_layout = makeTabLayout("NXT", "nxt_name", title_names[2], key_names[2])
    aew_layout = makeTabLayout("AEW", "aew_name", title_names[3], key_names[3])
    uk_layout = makeTabLayout("NXT UK", "uk_name", title_names[4], key_names[4])
    roh_layout = makeTabLayout("ROH", "roh_name", title_names[5], key_names[5])

    layout = [[sg.TabGroup([
        [sg.Tab('General', main_layout),
         sg.Tab('Raw', raw_layout, key='raw_tab'),
         sg.Tab('Smackdown', sd_layout, key='sd_tab'),
         sg.Tab('NXT', nxt_layout, key='nxt_tab', visible=False),
         sg.Tab('AEW', aew_layout, key='aew_tab', visible=False),
         sg.Tab('NXT UK', uk_layout, key='uk_tab', visible=False),
         sg.Tab('ROH', roh_layout, key='roh_tab', visible=False)
         ]])],
        [sg.Button("Generate"), sg.Exit()]
    ]

    # Create the window
    window = sg.Window("2K Universe Generator", layout)

    # Create the event loop
    while True:
        event, values = window.read()

        # a list of events to cause a refresh
        refresh = ["brands", "raw_name", "sd_name", "nxt_name", "aew_name", "uk_name", "roh_name"]

        # If Window closed or exit button pressed, end the program
        if event in (sg.WIN_CLOSED, 'Exit'):
            sys.exit()

        # If the Submit button is pressed
        elif event == "Generate":
            # Create variables for GUI inputs
            male_roster = values["-FILE_PATH-"]
            female_roster = values["-FILE_PATH2-"]
            brand_number = values["brands"]
            tag_teams = values["teams"]
            womens_tag_team_title = values["wTag"]
            second_midcard_title = values["mid2"]
            womens_midcard_title = values["wMid"]

            # list of the brand names
            brand_names = [values["raw_name"], values["sd_name"], values["nxt_name"],
                           values["aew_name"], values["uk_name"], values["roh_name"]]

            # list of the title names
            title_names = [
                [values["raw_world"], values["raw_mid"], values["raw_women"], values["raw_tag"], values["raw_mid2"], values["raw_womens_mid"]],
                [values["sd_world"], values["sd_mid"], values["sd_women"], values["sd_tag"], values["sd_mid2"], values["sd_womens_mid"]],
                [values["nxt_world"], values["nxt_mid"], values["nxt_women"], values["nxt_tag"], values["nxt_mid2"], values["nxt_womens_mid"]],
                [values["aew_world"], values["aew_mid"], values["aew_women"], values["aew_tag"], values["aew_mid2"], values["aew_womens_mid"]],
                [values["uk_world"], values["uk_mid"], values["uk_women"], values["uk_tag"], values["uk_mid2"], values["uk_womens_mid"]],
                [values["roh_world"], values["roh_mid"], values["roh_women"], values["roh_tag"], values["roh_mid2"], values["roh_womens_mid"]]
            ]

            # Create list of GUI inputs
            gui_inputs = [male_roster, female_roster, int(brand_number), int(tag_teams),
                          int(womens_tag_team_title), second_midcard_title, womens_midcard_title,
                          brand_names, title_names]

            # Check if brand number has exceeded the limit
            if int(brand_number) > 6: gui_inputs[2] = brand_number = 6

            # Check if tag team number has exceeded the limit
            if int(tag_teams) > 20: gui_inputs[3] = tag_teams = 20

            # Check if women's tag team number has exceeded the limit
            if int(womens_tag_team_title) > 10: gui_inputs[4] = womens_tag_team_title = 10

            # Check if a roster file is missing
            if not bool(male_roster) or not bool(female_roster):
                sg.PopupError(
                    "Missing Roster File", "One or both roster files has not been given. \n"
                                           "Please give both the men and women rosters.")
                continue

            # Check if a roster file is empty
            if len(getRoster(male_roster)) == 0 or len(getRoster(female_roster)) == 0:
                sg.PopupError("Empty Roster File", "One or both roster files is empty. \n"
                                                   "Please use a file with a list of participants")
                continue

            # Check if there are enough men/women to have as many tag teams as requested
            if len(getRoster(male_roster)) < int(tag_teams) * int(brand_number) * 2:
                sg.PopupError("Your men's roster is not big enough to support this many tag teams")
                continue

            if len(getRoster(female_roster)) < int(womens_tag_team_title) * int(brand_number) * 2:
                sg.PopupError("Your women's roster is not big enough to support this many tag teams.")
                continue
            # If any fields are blank, give an error message, and continue

            return gui_inputs

        # if a refresh event happens
        elif refresh.count(event) > 0:
            brand_number = values["brands"]

            window['raw_tab'].update(title=values["raw_name"])
            if brand_number == 1: window['sd_tab'].update(visible=False)
            else: window['sd_tab'].update(visible=True, title=values["sd_name"])

            if brand_number > 2: window['nxt_tab'].update(visible=True, title=values["nxt_name"])
            else: window['nxt_tab'].update(visible=False)

            if brand_number > 3: window['aew_tab'].update(visible=True, title=values["aew_name"])
            else: window['aew_tab'].update(visible=False)

            if brand_number > 4: window['uk_tab'].update(visible=True, title=values["uk_name"])
            else: window['uk_tab'].update(visible=False)

            if brand_number == 6: window['roh_tab'].update(visible=True, title=values["roh_name"])
            else: window['roh_tab'].update(visible=False)

    window.close()


# create excel spreadsheet of the draft
def createSpreadsheet(Brand, brand_number, draft_workbook, roster_sheet, tag_teams_sheet, champions_sheet,
                      divisions_sheet, feuds_sheet, women_tag_sheet, womens_tag_division):
    # write the rosters
    roster_sheet.cell(row=1, column=brand_number).value = Brand.name

    full_roster = Brand.men + Brand.women

    for row in range(len(full_roster)):
        roster_sheet.cell(row=row + 2, column=brand_number).value = full_roster[row]

    # write the tag teams

    full_teams = Brand.teams + Brand.women_teams
    tag_teams_sheet.cell(row=1, column=brand_number).value = Brand.name
    for row in range(len(full_teams)):
        tag_teams_sheet.cell(row=row + 2, column=brand_number).value = f"{full_teams[row][0]} & {full_teams[row][1]}"

    # write the champions

    # champion headers
    champions_sheet.cell(row=2, column=1).value = "World Champion"
    champions_sheet.cell(row=3, column=1).value = "Midcard Champion"
    champions_sheet.cell(row=4, column=1).value = "Tag Team Champions"
    champions_sheet.cell(row=5, column=1).value = "Women's Champion"
    champions_sheet.cell(row=6, column=1).value = "2nd Midcard Champion"
    champions_sheet.cell(row=7, column=1).value = "Women's Midcard Champion"

    # writing the cells
    champions_sheet.cell(row=1, column=brand_number + 1).value = Brand.name
    champions_sheet.cell(row=2, column=brand_number + 1).value = Brand.champions[0]
    champions_sheet.cell(row=3, column=brand_number + 1).value = Brand.champions[1]
    champions_sheet.cell(row=4, column=brand_number + 1).value = f"{Brand.champions[3][0]} & {Brand.champions[3][1]}"
    champions_sheet.cell(row=5, column=brand_number + 1).value = Brand.champions[2]
    champions_sheet.cell(row=6, column=brand_number + 1).value = Brand.champions[4]
    champions_sheet.cell(row=7, column=brand_number + 1).value = Brand.champions[5]

    # writing the women's tag champions
    if womens_tag_division:
        champions_sheet.cell(row=8, column=1).value = "Women's Tag Team Champions"
        champions_sheet.cell(row=8, column=2).value = f"{womens_tag_division[0][0]} & {womens_tag_division[0][1]}"

    # Writing the Divisions

    # Header
    divisions_sheet.cell(row=1, column=brand_number * 6 - 5).value = Brand.name
    divisions_sheet.merge_cells(start_row=1, start_column=brand_number * 6 - 5,
                                end_row=1, end_column=brand_number * 6)
    divisions_sheet.cell(row=2, column=brand_number * 6 - 5).value = f"{Brand.title_names[0]} Title"
    divisions_sheet.cell(row=2, column=brand_number * 6 - 4).value = f"{Brand.title_names[1]} Title"
    divisions_sheet.cell(row=2, column=brand_number * 6 - 3).value = f"{Brand.title_names[3]} Title"
    divisions_sheet.cell(row=2, column=brand_number * 6 - 2).value = f"{Brand.title_names[2]} Title"
    divisions_sheet.cell(row=2, column=brand_number * 6 - 1).value = f"{Brand.title_names[4]} Title"
    divisions_sheet.cell(row=2, column=brand_number * 6).value = f"{Brand.title_names[5]} Title"

    # Body

    # World Divisions
    for row in range(len(Brand.divisions[0])):
        divisions_sheet.cell(row=row + 3, column=brand_number * 6 - 5).value = Brand.divisions[0][row]

    # Midcard Divisions
    for row in range(len(Brand.divisions[1])):
        divisions_sheet.cell(row=row + 3, column=brand_number * 6 - 4).value = Brand.divisions[1][row]

    # Tag Team Division
    for row in range(len(Brand.divisions[3])):
        divisions_sheet.cell(row=row + 3,
                             column=brand_number * 6 - 3).value = f"{Brand.divisions[3][row][0]} & {Brand.divisions[3][row][1]}"

    # Women's Division
    for row in range(len(Brand.divisions[2])):
        divisions_sheet.cell(row=row + 3, column=brand_number * 6 - 2).value = Brand.divisions[2][row]

    # 2nd Midcard Division
    if Brand.divisions[4]:
        for row in range(len(Brand.divisions[4])):
            divisions_sheet.cell(row=row + 3, column=brand_number * 6 - 1).value = Brand.divisions[4][row]

    # Women's Midcard Division
    if Brand.divisions[5]:
        for row in range(len(Brand.divisions[5])):
            divisions_sheet.cell(row=row + 3, column=brand_number * 6).value = Brand.divisions[5][row]

    # Women's Tag Division
    if womens_tag_division:
        women_tag_sheet.cell(row=1, column=1).value = "Women's Tag Team Division"
        for row in range(len(womens_tag_division)):
            women_tag_sheet.cell(row=row + 2,
                                 column=1).value = f"{womens_tag_division[row][0]} & {womens_tag_division[row][1]}"

    # Writing The Feuds
    feuds_sheet.cell(row=1, column=brand_number).value = Brand.name
    for row in range(len(Brand.feuds)):
        feuds_sheet.cell(row=row + 2, column=brand_number).value = Brand.feuds[row]

    # Color in the Brand Headers

    # set colors to each brand
    raw_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sd_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    nxt_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    aew_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    uk_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    roh_fill = PatternFill(start_color='775973', end_color='775973', fill_type='solid')

    # raw
    roster_sheet['A1'].fill = raw_fill
    tag_teams_sheet['A1'].fill = raw_fill
    champions_sheet['B1'].fill = raw_fill
    divisions_sheet['A1'].fill = raw_fill
    feuds_sheet['A1'].fill = raw_fill

    # smackdown
    if brand_number > 1:
        roster_sheet['B1'].fill = sd_fill
        tag_teams_sheet['B1'].fill = sd_fill
        champions_sheet['C1'].fill = sd_fill
        divisions_sheet['G1'].fill = sd_fill
        feuds_sheet['B1'].fill = sd_fill

    # nxt
    if brand_number > 2:
        roster_sheet['C1'].fill = nxt_fill
        tag_teams_sheet['C1'].fill = nxt_fill
        champions_sheet['D1'].fill = nxt_fill
        divisions_sheet['M1'].fill = nxt_fill
        feuds_sheet['C1'].fill = nxt_fill

    # aew
    if brand_number > 3:
        roster_sheet['D1'].fill = aew_fill
        tag_teams_sheet['D1'].fill = aew_fill
        champions_sheet['E1'].fill = aew_fill
        divisions_sheet['S1'].fill = aew_fill
        feuds_sheet['D1'].fill = aew_fill

    # uk
    if brand_number > 4:
        roster_sheet['E1'].fill = uk_fill
        tag_teams_sheet['E1'].fill = uk_fill
        champions_sheet['F1'].fill = uk_fill
        divisions_sheet['Y1'].fill = uk_fill
        feuds_sheet['E1'].fill = uk_fill

    # roh
    if brand_number == 6:
        roster_sheet['F1'].fill = roh_fill
        tag_teams_sheet['F1'].fill = roh_fill
        champions_sheet['G1'].fill = roh_fill
        divisions_sheet['AE1'].fill = roh_fill
        feuds_sheet['F1'].fill = roh_fill

    draft_workbook.save("Draft.xlsx")


# add formating to the excel spreadsheet
def formatSpreadsheet(cell_range, style):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in cell_range:
        for c in cell:
            if style == "Header":
                c.font = Font(name="Arial", size=8, bold=True)
                c.alignment = Alignment(horizontal='center')
            elif style == "Body":
                c.font = Font(name="Arial", size=9, bold=False)
            c.border = thin_border


def developBrand(Brand, team_number, women_tag_number, second_midcard, women_midcard, draft_file):
    Brand.createTeams(team_number)
    Brand.createWomenTeams(women_tag_number)
    Brand.assignChampions(Brand.teams, second_midcard, women_midcard)
    Brand.assignDivisions(Brand.teams, Brand.champions, second_midcard, women_midcard)
    Brand.assignFeuds(Brand.divisions)
    makeTxtFile(draft_file, Brand.name, Brand.men, Brand.women, Brand.teams, Brand.champions, Brand.divisions,
                second_midcard, women_midcard, Brand.feuds, Brand.title_names)


def createWomenTagDivision(women_tag_number, brands, draft_file):
    if women_tag_number == 0: return

    womens_tag_division = []
    brand_number = len(brands)
    match brand_number:
        case 1:
            womens_tag_division = brands[0].women_teams
        case 2:
            womens_tag_division = brands[0].women_teams + brands[1].women_teams
        case 3:
            womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams
        case 4:
            womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + brands[
                3].women_teams
        case 5:
            womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + brands[
                3].women_teams + brands[4].women_teams
        case 6:
            womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + brands[
                3].women_teams + brands[4].women_teams + brands[5].women_teams

    random.shuffle(womens_tag_division)
    womens_tag_champions = womens_tag_division[0]
    writeWomenTag(draft_file, womens_tag_division, womens_tag_champions)
    return womens_tag_division


def main():
    inputs = gui()

    # create variables from inputs
    male_roster_file, female_roster_file = inputs[0], inputs[1]  # roster inputs
    brand_number, team_number = inputs[2], inputs[3]  # brand inputs
    women_tag_number, second_midcard, women_midcard = inputs[4], inputs[5], inputs[6]  # optional division inputs
    brand_names, title_names = inputs[7], inputs[8] # name inputs

    # get the rosters
    male_roster, female_roster = getRoster(male_roster_file), getRoster(female_roster_file)

    # create the draft text file
    draft_file = open("draft.txt", 'w')
    draft_file.close()
    draft_file = open("draft.txt", 'a')

    # create the draft spreadsheet
    draft_spreadsheet = xl.Workbook()
    rosters_sheet = draft_spreadsheet.create_sheet("Rosters")
    tag_teams_sheet = draft_spreadsheet.create_sheet("Tag Teams")
    champions_sheet = draft_spreadsheet.create_sheet("Champions")
    divisions_sheet = draft_spreadsheet.create_sheet("Divisions")
    feuds_sheet = draft_spreadsheet.create_sheet("Rivalries")
    women_tag_sheet = draft_spreadsheet.create_sheet("Women's Tag Division")

    # format the spreadsheet

    # draft the rosters
    men, women = draftRoster(male_roster, brand_number), draftRoster(female_roster, brand_number)

    # create the brands
    brands = []  # list of the brands

    # create first brand
    Raw = Brand(brand_names[0], men[0], women[0], title_names[0])
    developBrand(Raw, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
    brands.append(Raw)

    # create second brand
    if brand_number > 1:
        Smackdown = Brand(brand_names[1], men[1], women[1], title_names[1])
        developBrand(Smackdown, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
        brands.append(Smackdown)

    # create third brand
    if brand_number > 2:
        NXT = Brand(brand_names[2], men[2], women[2], title_names[2])
        developBrand(NXT, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
        brands.append(NXT)

    # create fourth brand
    if brand_number > 3:
        AEW = Brand(brand_names[3], men[3], women[3], title_names[3])
        developBrand(AEW, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
        brands.append(AEW)

    # create fifth brand
    if brand_number > 4:
        UK = Brand(brand_names[4], men[4], women[4], title_names[4])
        developBrand(UK, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
        brands.append(UK)

    # create sixth brand
    if brand_number > 5:
        ROH = Brand(brand_names[5], men[5], women[5], title_names[5])
        developBrand(ROH, team_number, women_tag_number, second_midcard, women_midcard, draft_file)
        brands.append(ROH)

    # set womens tag team champions if division exists
    womens_tag_division = createWomenTagDivision(women_tag_number, brands, draft_file)

    # create spreadsheet
    i = 1  # iterator
    for b in brands:
        createSpreadsheet(b, i, draft_spreadsheet, rosters_sheet, tag_teams_sheet, champions_sheet, divisions_sheet,
                          feuds_sheet,
                          women_tag_sheet, womens_tag_division)
        i += 1

    draft_file.close()  # close the file handler

    headers = [rosters_sheet['A1:Z1'], tag_teams_sheet['A1:Z1'], champions_sheet['A1:Z1'], champions_sheet['A2:A20'],
               divisions_sheet['A1:BA2'], women_tag_sheet['A1:Z1'], feuds_sheet['A1:Z1']]
    bodies = [rosters_sheet['A2:K200'], tag_teams_sheet['A2:K50'], champions_sheet['B2:K50'],
              divisions_sheet['A3:BA200'], women_tag_sheet['A2:Z50'], feuds_sheet['A2:Z50']]

    for h in headers:
        formatSpreadsheet(h, "Header")

    for b in bodies:
        formatSpreadsheet(b, "Body")
    std = draft_spreadsheet['Sheet']
    draft_spreadsheet.remove(std)

    for col in range(1, divisions_sheet.max_column + 1):
        rosters_sheet.column_dimensions[get_column_letter(col)].width = 25
        tag_teams_sheet.column_dimensions[get_column_letter(col)].width = 30
        champions_sheet.column_dimensions[get_column_letter(col)].width = 25
        divisions_sheet.column_dimensions[get_column_letter(col)].width = 25
        women_tag_sheet.column_dimensions[get_column_letter(col)].width = 25
        feuds_sheet.column_dimensions[get_column_letter(col)].width = 30

    draft_spreadsheet.save("Draft.xlsx")
    sg.PopupOK("Success", "Your Universe has been successfully generated.\n"
                          "Look for the Draft.txt and Draft.xlsx files in the program's directory for the results")


main()
