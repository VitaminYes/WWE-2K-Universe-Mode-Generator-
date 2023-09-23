import random, os
import PySimpleGUI as sg
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
from Brand import *


# additional functions

# create the rosters
def getRoster(roster_file):
    with open(roster_file) as f:
        roster = f.readlines()
    return roster


# draft the brands
def draftRoster(roster, brands):
    raw_roster, smackdown_roster, nxt_roster, aew_roster, nxt_uk_roster, roh_roster = [], [], [], [], [], []  # Roster lists
    b = 1  # Brand Counter
    random.shuffle(roster)  # Shuffle the roster order

    # Draft the rosters

    for r in roster:
        match b:
            case 1:
                raw_roster.append(r[:-1])
                if brands > 1: b = 2
                continue
            case 2:
                smackdown_roster.append(r[:-1])
                if brands == 2:
                    b = 1
                else:
                    b = 3
                continue
            case 3:
                nxt_roster.append(r[:-1])
                if brands == 3:
                    b = 1
                else:
                    b = 4
                continue
            case 4:
                aew_roster.append(r[:-1])
                if brands == 4:
                    b = 1
                else:
                    b = 5
                continue
            case 5:
                nxt_uk_roster.append(r[:-1])
                if brands == 5:
                    b = 1
                else:
                    b = 6
                continue
            case 6:
                roh_roster.append(r[:-1]);
                b = 1;
                continue

    return [raw_roster, smackdown_roster, nxt_roster, aew_roster, nxt_uk_roster, roh_roster]


def addTeams(roster, teams):
    for t in teams: roster.append(t[0]); roster.append(t[1])
    return roster


def makeTabLayout(brand_name, brand_key, title_names, key_names):
    brand_layout = [[sg.Text("Brand Name: "), sg.Input(key=brand_key, default_text=brand_name, enable_events=True)],
                    [sg.Text("World Title Name: "), sg.Input(key=key_names[0], default_text=title_names[0])],
                    [sg.Text("Midcard Title Name: "), sg.Input(key=key_names[1], default_text=title_names[1])],
                    [sg.Text("Women's Title Name: "), sg.Input(key=key_names[2], default_text=title_names[2])],
                    [sg.Text("Tag Team Title Name: "), sg.Input(key=key_names[3], default_text=title_names[3])],
                    [sg.Text("Second Midcard Title Name: "), sg.Input(key=key_names[4], default_text=title_names[4])],
                    [sg.Text("Women's Midcard Title Name: "), sg.Input(key=key_names[5], default_text=title_names[5])]
                    ]
    return brand_layout


# add the info of a brand to the text file
def makeTxtFile(f, Brand, second_midcard_division, womens_midcard_division, file_type, team_roster):
    if file_type == "Spreadsheet": return

    title_names = Brand.title_names

    # Header
    f.write(f"{Brand.name}\n")
    f.write("=" * 30 + "\n\n")

    if not team_roster: full_mens_roster = addTeams(Brand.men, Brand.teams)
    else: full_mens_roster = Brand.men

    # Men's Roster
    f.write("Men's Roster\n")
    f.write("=" * 30 + "\n\n")
    for m in full_mens_roster: f.write(f"{m}\n")
    f.write("\n")

    # Women's Roster
    f.write("Women's Roster\n")
    f.write("=" * 30 + "\n")
    for w in Brand.women: f.write(f"{w}\n")
    f.write("\n")

    # Tag Team
    f.write("Tag Teams\n")
    f.write("=" * 30 + "\n")
    for t in Brand.teams:
        if team_roster: f.write(f"{t}\n")
        else: f.write(f"{t[0]} & {t[1]}\n")
    f.write("\n")

    # Champions
    f.write("Champions \n")
    f.write("=" * 30 + "\n")
    f.write(f"{title_names[0]} Champion: {Brand.champions[0]}\n") # World Title
    f.write(f"{title_names[1]} Champion: {Brand.champions[1]}\n")  # Midcard Title
    if second_midcard_division: f.write(f"{title_names[4]} Champion: {Brand.champions[4]}\n") # 2nd Midcard Title
    if team_roster: f.write(f"{title_names[3]}: {Brand.champions[3]}\n") # Tag Title
    else: f.write(f"{title_names[3]}: {Brand.champions[3][0]} & {Brand.champions[3][1]}\n")
    f.write(f"{title_names[2]} Champion: {Brand.champions[2]}\n") # Women's Title
    if womens_midcard_division: f.write(f"{title_names[5]} Champion: {Brand.champions[5]}") # Women's Midcard Title
    f.write("\n\n")

    # Divisions
    f.write("Divisions\n")
    f.write("=" * 30 + "\n\n")

    # World Title
    f.write(f"{title_names[0]} Title\n")
    f.write("=" * 30 + "\n")
    for d in Brand.divisions[0]: f.write(f"{d}\n")
    f.write("\n\n")

    # Midcard Title
    f.write(f"{title_names[1]} Title\n")
    f.write("=" * 30 + "\n")
    for d in Brand.divisions[1]: f.write(f"{d}\n")
    f.write("\n\n")

    # 2nd Midcard Title
    if second_midcard_division:
        f.write(f"{title_names[4]} Title\n")
        f.write("=" * 30 + "\n")
        for d in Brand.divisions[4]: f.write(f"{d}\n")
        f.write("\n\n")

    # Tag Team Title
    f.write(f"{title_names[3]} Title\n")
    f.write("=" * 30 + "\n")
    for d in Brand.divisions[3]:
        if team_roster: f.write(f"{d}\n")
        else: f.write(f"{d[0]} & {d[1]}\n")
    f.write("\n\n")

    # Women's Title
    f.write(f"{title_names[2]} Title\n")
    f.write("=" * 30 + "\n")
    for d in Brand.divisions[2]: f.write(f"{d}\n")
    f.write("\n\n")

    # Women's Midcard Title
    if womens_midcard_division:
        f.write(f"{title_names[5]} Title\n")
        f.write("=" * 30 + "\n")
        for d in Brand.divisions[5]: f.write(f"{d}\n")
        f.write("\n\n")

    # Feuds
    f.write("Rivalries\n");
    f.write("=" * 30 + "\n")
    for r in Brand.feuds: f.write(f"{r}\n")
    f.write("\n\n")


# add the women's tag division to the text file
def writeWomenTag(f, womens_teams, womens_tag_champions, women_team_roster):
    f.write("\n\n")
    f.write("Women's Tag Division\n")
    f.write("=" * 30)
    f.write("\n\n")  # create header

    if women_team_roster:
        f.write(f"Women's Tag Team Champions: {womens_tag_champions}\n\n")
        for w in womens_teams[1:]: f.write(f"{w}\n")
    else:
        f.write(f"Women's Tag Team Champions: {womens_tag_champions[0]} & {womens_tag_champions[1]}\n\n")
        for w in womens_teams[1:]: f.write(f"{w[0]} & {w[1]}\n")


# create excel spreadsheet of the draft
def createSpreadsheet(Brand, brand_number, draft_workbook, roster_sheet, tag_teams_sheet, champions_sheet,
        divisions_sheet, feuds_sheet, women_tag_sheet, womens_tag_division, file_type, team_roster, women_team_roster):

    tn = Brand.title_names

    # write the rosters
    roster_sheet.cell(row=1, column=brand_number).value = Brand.name

    if team_roster:
        full_roster = Brand.men + Brand.women
        full_roster.sort()
        full_roster = full_roster + Brand.teams
    else: full_roster = addTeams(Brand.men + Brand.women, Brand.teams)
    full_roster.sort()

    # setting the number of titles
    title_number = 4
    if Brand.champions[4]: title_number += 1
    if Brand.champions[5]: title_number += 1

    for row in range(len(full_roster)): roster_sheet.cell(row=row + 2, column=brand_number).value = full_roster[row]

    # write the tag teams
    if not women_team_roster: full_teams = Brand.teams + Brand.women_teams
    else: full_teams = Brand.teams
    tag_teams_sheet.cell(row=1, column=brand_number).value = Brand.name
    for row in range(len(full_teams)):
        if team_roster: tag_teams_sheet.cell(row=row + 2, column=brand_number).value = f"{full_teams[row]}"
        else: tag_teams_sheet.cell(row=row + 2, column=brand_number).value = f"{full_teams[row][0]} & {full_teams[row][1]}"

    if women_team_roster:
        for row in range(len(Brand.women_teams)):
            tag_teams_sheet.cell(row=row + len(Brand.teams) + 2, column=brand_number).value = Brand.women_teams[row]

    # write the champions

    # champion headers
    champions_sheet.cell(row=2, column=1).value = "World Champion"
    champions_sheet.cell(row=3, column=1).value = "Midcard Champion"
    champions_sheet.cell(row=4, column=1).value = "Tag Team Champions"
    champions_sheet.cell(row=5, column=1).value = "Women's Champion"
    if Brand.champions[4]: champions_sheet.cell(row=6, column=1).value = "2nd Midcard Champion"
    if Brand.champions[5]: champions_sheet.cell(row=7, column=1).value = "Women's Midcard Champion"

    # writing the cells
    champions_sheet.cell(row=1, column=brand_number + 1).value = Brand.name
    champions_sheet.cell(row=2, column=brand_number + 1).value = Brand.champions[0] # World Champion
    champions_sheet.cell(row=3, column=brand_number + 1).value = Brand.champions[1] # Midcard Champion
    if team_roster: champions_sheet.cell(row=4, column=brand_number + 1).value = f"{Brand.champions[3]}"
    else: champions_sheet.cell(row=4, column=brand_number + 1).value = f"{Brand.champions[3][0]} & {Brand.champions[3][1]}" # Tag Team Champion
    champions_sheet.cell(row=5, column=brand_number + 1).value = Brand.champions[2]
    champions_sheet.cell(row=6, column=brand_number + 1).value = Brand.champions[4]
    champions_sheet.cell(row=7, column=brand_number + 1).value = Brand.champions[5]

    # writing the women's tag champions
    if womens_tag_division:
        champions_sheet.cell(row=8, column=1).value = "Women's Tag Team Champions"
        if women_team_roster: champions_sheet.cell(row=8, column=2).value = f"{womens_tag_division[0]}"
        else: champions_sheet.cell(row=8, column=2).value = f"{womens_tag_division[0][0]} & {womens_tag_division[0][1]}"

    # writing the divisions


    # Header
    divisions_sheet.cell(row=1, column=brand_number * title_number - (title_number - 1)).value = Brand.name
    divisions_sheet.merge_cells(start_row=1, start_column=brand_number * title_number - (title_number - 1),
                                end_row=1, end_column=brand_number * title_number)

    divisions_sheet.cell(row=2, column=brand_number * title_number - (title_number - 1)).value = f"{tn[0]} Title"
    divisions_sheet.cell(row=2, column=brand_number * title_number - (title_number - 2)).value = f"{tn[1]} Title"
    divisions_sheet.cell(row=2, column=brand_number * title_number - (title_number - 3)).value = f"{tn[3]} Title"
    divisions_sheet.cell(row=2, column=brand_number * title_number - (title_number - 4)).value = f"{tn[2]} Title"
    if Brand.champions[4]: divisions_sheet.cell(row=2, column=brand_number * title_number - (title_number - 5)).value = f"{tn[4]} Title"
    if Brand.champions[5]: divisions_sheet.cell(row=2,column=brand_number * title_number).value = f"{tn[5]} Title"

    # Body

    # World Divisions
    for row in range(len(Brand.divisions[0])):
        divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 1)).value = Brand.divisions[0][row]

    # Midcard Divisions
    for row in range(len(Brand.divisions[1])):
        divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 2)).value = Brand.divisions[1][row]

    # Tag Team Division
    for row in range(len(Brand.divisions[3])):
        if team_roster:
            divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 3)).value = \
                f"{Brand.divisions[3][row]}"
        else:
            divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 3)).value = \
                f"{Brand.divisions[3][row][0]} & {Brand.divisions[3][row][1]}"

    # Women's Division
    for row in range(len(Brand.divisions[2])):
        divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 4)).value = \
            Brand.divisions[2][row]

    # 2nd Midcard Division
    if Brand.divisions[4]:
        for row in range(len(Brand.divisions[4])):
            divisions_sheet.cell(row=row + 3, column=brand_number * title_number - (title_number - 5)).value = \
                Brand.divisions[4][row]

    # Women's Midcard Division
    if Brand.divisions[5]:
        for row in range(len(Brand.divisions[5])):
            divisions_sheet.cell(row=row + 3, column=brand_number * title_number).value = Brand.divisions[5][row]

    # Women's Tag Division
    if womens_tag_division:
        women_tag_sheet.cell(row=1, column=1).value = "Women's Tag Team Division"
        for row in range(len(womens_tag_division)):
            if women_team_roster: women_tag_sheet.cell(row=row + 2, column=1).value = f"{womens_tag_division[row]}"
            else: women_tag_sheet.cell(row=row + 2, column=1).value = f"{womens_tag_division[row][0]} & {womens_tag_division[row][1]}"

    # Writing The Feuds
    feuds_sheet.cell(row=1, column=brand_number).value = Brand.name
    for row in range(len(Brand.feuds)):
        feuds_sheet.cell(row=row + 2, column=brand_number).value = Brand.feuds[row]

    # Color in the Brand Headers

    brand_colors = [
        PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),
        PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        PatternFill(start_color='775973', end_color='775973', fill_type='solid')
    ]

    # color in headings
    for b in range(1, brand_number + 1):
        division_header = b * title_number - (title_number - 1)
        roster_sheet.cell(row=1, column=b).fill = brand_colors[b - 1]
        tag_teams_sheet.cell(row=1, column=b).fill = brand_colors[b - 1]
        champions_sheet.cell(row=1, column=b + 1).fill = brand_colors[b - 1]
        divisions_sheet.cell(row=1, column=division_header).fill = brand_colors[b - 1]
        feuds_sheet.cell(row=1, column=b).fill = brand_colors[b - 1]

    if file_type == "Spreadsheet": draft_workbook.save("Draft.xlsx")


# add formating to the excel spreadsheet
def formatSpreadsheet(cell_range, style):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in cell_range:
        for c in cell:
            if style == "Header":
                c.font = Font(name="Arial", size=8, bold=True)
                c.alignment = Alignment(horizontal='center')
            elif style == "Body":
                c.font = Font(name="Arial", size=9, bold=False)
            c.border = thin_border


def developBrand(Brand, team_number, women_tag_number, second_midcard, women_midcard, draft_file, file_type,
                 teams, women_teams):
    # add tag teams
    if teams: Brand.teams = teams
    else: Brand.createTeams(team_number)

    # add womens tag teams
    if women_teams: Brand.women_teams = women_teams
    else: Brand.createWomenTeams(women_tag_number)

    Brand.assignChampions(Brand.teams, second_midcard, women_midcard, teams)
    Brand.assignDivisions(Brand.teams, Brand.champions, second_midcard, women_midcard, teams)
    Brand.assignFeuds(Brand.divisions, teams)
    if file_type == "Text":
        makeTxtFile(draft_file, Brand, second_midcard, women_midcard, "Text", teams)


def createWomenTagDivision(women_tag_number, brands, draft_file, file_type, women_team_roster):
    if women_tag_number == 0: return

    womens_tag_division = []
    brand_number = len(brands)
    match brand_number:
        case 1: womens_tag_division = brands[0].women_teams
        case 2: womens_tag_division = brands[0].women_teams + brands[1].women_teams
        case 3: womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams
        case 4: womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + \
                    brands[3].women_teams
        case 5: womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + \
                    brands[3].women_teams + brands[4].women_teams
        case 6: womens_tag_division = brands[0].women_teams + brands[1].women_teams + brands[2].women_teams + \
                    brands[3].women_teams + brands[4].women_teams + brands[5].women_teams

    random.shuffle(womens_tag_division)
    womens_tag_champions = womens_tag_division[0]
    if file_type == "Text": writeWomenTag(draft_file, womens_tag_division, womens_tag_champions, women_team_roster)
    return womens_tag_division
